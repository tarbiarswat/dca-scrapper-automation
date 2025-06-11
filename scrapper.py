#Adds status column in excel output for exact matches with custom column autofit width, highlights cell color for a matched row and generates a log which captures timestamp and action points or the matching log

import time
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def clean_text(text):
    return text.replace("\xa0", " ").strip()

def log_to_file(message):
    timestamp = datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
    with open("search_log.txt", "a", encoding="utf-8") as f:
        f.write(f"{timestamp} {message}\n")

def process_single_name(first_name, last_name):
    options = Options()
    options.add_argument("--window-size=1000,800")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    FIELDS = [
        "First Name", "Middle Name", "Last Name", "License Number", "License Type",
        "License Status", "Expiration Date", "Secondary Status", "City", "State", "County", "Zip", "Match Status"
    ]

    results = []
    match_found = False

    try:
        driver.get("https://search.dca.ca.gov/?BD=19&TP=DC")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "firstName"))).send_keys(first_name)
        driver.find_element(By.ID, "lastName").send_keys(last_name)
        driver.find_element(By.ID, "licenseType").send_keys("Certified Public Accountant")
        driver.find_element(By.ID, "srchSubmitHome").click()

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'article.post')))
        except TimeoutException:
            print(f"❌ No results returned for: {first_name} {last_name}")
            log_to_file(f"No results returned for: {first_name} {last_name}")
            return [{
                "First Name": first_name,
                "Middle Name": " - ",
                "Last Name": last_name,
                "License Number": " - ",
                "License Type": " - ",
                "License Status": " - ",
                "Expiration Date": " - ",
                "Secondary Status": " - ",
                "City": " - ",
                "State": " - ",
                "County": " - ",
                "Zip": " - ",
                "Match Status": "Not Found"
            }]

        articles = driver.find_elements(By.CSS_SELECTOR, 'article.post')
        print(f"\n🔍 Found {len(articles)} total records for: {first_name} {last_name}")

        for idx, article in enumerate(articles):
            data = {field: "" for field in FIELDS}
            try:
                li_items = article.find_elements(By.CSS_SELECTOR, "ul.actions li")
                for li in li_items:
                    html = li.get_attribute("innerHTML")
                    soup = BeautifulSoup(html, "html.parser")
                    text = soup.get_text(separator=" ", strip=True)

                    if "<h3>" in html:
                        try:
                            last, first_middle = text.split(",", 1)
                            name_parts = first_middle.strip().split()
                            data["Last Name"] = clean_text(last)
                            data["First Name"] = name_parts[0] if name_parts else ""
                            data["Middle Name"] = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
                        except ValueError:
                            data["Last Name"] = text
                    elif ":" in text:
                        key, value = [clean_text(x) for x in text.split(":", 1)]
                        if key in data:
                            data[key] = value

                print(f"📄 Record #{idx}")
                for k in FIELDS[:-1]:
                    print(f"   {k}: {data.get(k, '')}")
                print("")

                if (data["State"].strip().lower() == "california" and
                    data["First Name"].strip().lower() == first_name.strip().lower() and
                    data["Last Name"].strip().lower() == last_name.strip().lower()):
                    data["Match Status"] = "Matched"
                    results.append(data)
                    match_found = True
                    log_to_file(f"Record #{idx}: ✅ Match found for {first_name} {last_name}")
                else:
                    log_to_file(f"Record #{idx}: ⛔ Not matched for {first_name} {last_name}")

            except Exception as e:
                print(f"⚠️ Error parsing record #{idx}: {e}")
                log_to_file(f"Record #{idx}: ⚠️ Error for {first_name} {last_name}: {e}")

    finally:
        driver.quit()

        if not match_found:
            print(f"❌ No exact match found in California for: {first_name} {last_name}")
            log_to_file(f"No exact match in California for: {first_name} {last_name}")
            results.append({
                "First Name": first_name,
                "Middle Name": " - ",
                "Last Name": last_name,
                "License Number": " - ",
                "License Type": " - ",
                "License Status": " - ",
                "Expiration Date": " - ",
                "Secondary Status": " - ",
                "City": " - ",
                "State": " - ",
                "County": " - ",
                "Zip": " - ",
                "Match Status": "Not Found"
            })

        return results

def main():
    df = pd.read_excel("names.xlsx", skiprows=2)
    all_results = []
    for _, row in df.iterrows():
        first_name = str(row[0]).strip()
        last_name = str(row[1]).strip()
        if first_name and last_name:
            all_results.extend(process_single_name(first_name, last_name))

    output_path = "california_matches.xlsx"
    df_result = pd.DataFrame(all_results)
    df_result.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                continue
        ws.column_dimensions[column_letter].width = max_length + 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status_cell = row[-1]  # Last column = "Match Status"
        if status_cell.value and str(status_cell.value).strip().lower() == "matched":
            for cell in row:
                cell.fill = green_fill

    wb.save(output_path)
    print(f"\n✅ Done! Results saved to '{output_path}' with autofit columns and highlight.")
    log_to_file(f"✅ Final Excel saved: {output_path}")

if __name__ == "__main__":
    main()
